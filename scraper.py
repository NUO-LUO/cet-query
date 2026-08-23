# -*- coding: utf-8 -*-
"""
CET 四六级成绩批量查询爬虫
===========================
使用方法:
  - 本地运行: python scraper.py
  - GitHub Actions 定时运行: 见 .github/workflows/scrape.yml
  - 结果写入 scores.json,配合 index.html 部署到 GitHub Pages

说明:
  - 成绩发布前运行,首个查询会返回 code=403 并正常退出(exit 0)
  - 身份证号在写入 scores.json 前已脱敏(前4后4),名单不进公开仓库
  - 支持断点续跑:已查询过(queried=true)的学生会自动跳过
  - 请求间隔 2 秒 + 每 50 人冷却 10 秒,730 人约需 35 分钟
"""

import json
import os
import subprocess
import sys
import time
from datetime import datetime

# 强制实时输出，避免 GitHub Actions 日志缓冲看不到进度
sys.stdout.reconfigure(line_buffering=True)

try:
    import requests
except ImportError:
    print("缺少 requests 库,正在安装...")
    os.system(f"{sys.executable} -m pip install requests -i https://pypi.tuna.tsinghua.edu.cn/simple")
    import requests

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl 库,正在安装...")
    os.system(f"{sys.executable} -m pip install openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple")
    import openpyxl


# ==================== 配置 ====================
# 逆向官网 JS 得到的真实 API 地址(无需 Cookie)
API_URL = "https://appquery.neea.edu.cn/latest/results/cet"
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "account", "福清华侨中学学生名单.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scores.json")
REQUEST_INTERVAL = 2  # 请求间隔(秒)，防限流
# ================================================


def read_students():
    """从 Excel 读取所有学生信息"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    students = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[1] and row[2]:
                students.append({
                    "class": str(row[0]).strip(),
                    "name": str(row[1]).strip(),
                    "id": str(row[2]).strip()
                })
    wb.close()
    return students


def mask_id(id_num):
    """脱敏身份证号:保留前4后4,中间用 * 遮挡"""
    if len(id_num) <= 8:
        return id_num
    return id_num[:4] + "********" + id_num[-4:]


def load_existing():
    """读取已有 scores.json,返回 {class|name: item},用于断点续跑"""
    existing = {}
    if os.path.exists(OUTPUT_FILE):
        try:
            with open(OUTPUT_FILE, encoding="utf-8") as f:
                old = json.load(f)
            for item in old.get("data", []):
                if item.get("queried"):
                    key = f"{item.get('class')}|{item.get('name')}"
                    existing[key] = item
        except Exception:
            pass
    return existing


def query_score(session, name, id_number, km=1):
    """
    查询单个学生的 CET 成绩
    km: 1=四级, 2=六级
    返回: (成绩 dict 或 None, 是否得到可确认的官方响应)

    第二个返回值为 False 表示网络、限流或服务端异常；这类学生会在下次
    运行时重试，而不是被错误地标记为已查询。
    """
    params = {
        "km": km,
        "xm": name,
        "no": id_number,
        "source": "mb"
    }
    for attempt in range(1, 4):
        try:
            resp = session.get(API_URL, params=params, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                code = data.get("code")
                if code == 0:
                    return {
                        "total": data.get("score"),
                        "listening": data.get("sco_lc"),
                        "reading": data.get("sco_rd"),
                        # 官网成绩详情标为“写作和翻译”。
                        "writing": data.get("sco_wt"),
                        "ticket": data.get("zkzh", ""),
                        "level": "四级" if km == 1 else "六级"
                    }, True
                if code == 403:
                    msg = data.get("msg", "")
                    print(f"\n[暂停] {msg}")
                    print("成绩尚未发布,脚本正常退出,等待下次调度。")
                    sys.exit(0)

                # 姓名/证件号不匹配、未参加考试等，属于可确认的结果。
                return None, True

            print(f"\n  [警告] HTTP {resp.status_code}，第 {attempt}/3 次请求失败")
        except requests.exceptions.Timeout:
            print(f"\n  [超时] 第 {attempt}/3 次请求超时")
        except Exception as e:
            print(f"\n  [错误] 第 {attempt}/3 次请求异常: {e}")

        if attempt < 3:
            time.sleep(5)

    return None, False


def save_results(results):
    """保存查询结果到 scores.json(身份证号已脱敏)"""
    found = sum(1 for r in results if r.get("cet4") or r.get("cet6"))
    output = {
        "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_students": len(results),
        "queried_count": found,
        "data": results
    }
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)


def push_progress(count, total):
    """在 GitHub Actions 环境中将进度 push 回仓库，实现真正的断点续跑"""
    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        return  # 本地运行，跳过
    try:
        repo = os.environ.get("GITHUB_REPOSITORY", "NUO-LUO/cet-query")
        branch = os.environ.get("GITHUB_REF_NAME", "master")
        subprocess.run(
            ["git", "config", "user.name", "github-actions[bot]"], check=False,
            capture_output=True)
        subprocess.run(
            ["git", "config", "user.email", "github-actions[bot]@users.noreply.github.com"],
            check=False, capture_output=True)
        subprocess.run(
            ["git", "remote", "set-url", "origin",
             f"https://x-access-token:{token}@github.com/{repo}.git"],
            check=False, capture_output=True)
        subprocess.run(["git", "add", "scores.json"], check=False, capture_output=True)
        r = subprocess.run(
            ["git", "commit", "-m", f"chore: 查询进度 {count}/{total}"],
            check=False, capture_output=True, text=True)
        if "nothing to commit" not in r.stdout and "no changes" not in r.stdout.lower():
            subprocess.run(
                ["git", "push", "origin", f"HEAD:{branch}"],
                check=False, capture_output=True)
            print(f"  >>> 进度已 push 到仓库 ({count}/{total})", flush=True)
    except Exception as e:
        print(f"  [警告] push 进度失败: {e}", flush=True)


def main():
    print("=" * 50)
    print("  CET 四六级成绩批量查询工具")
    print("=" * 50)

    # 读取学生列表
    print(f"\n[1/3] 读取学生名单...")
    students = read_students()
    total = len(students)
    print(f"  共 {total} 名学生")

    short_ids = sum(1 for s in students if len(s["id"]) < 15)
    if short_ids > 0:
        print(f"  注意: {short_ids} 人身份证号不完整,将标记为 id_incomplete")

    # 断点续跑:加载已查询过的学生
    print(f"\n[2/3] 检查已有结果...")
    existing = load_existing()
    if existing:
        print(f"  已查询过 {len(existing)} 人,将跳过")
        if len(existing) == total:
            print("  全部学生均已查询，无需重复请求。")
            return

    # 创建 session
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/126.0.0.0 Safari/537.36",
        "Accept": "*/*",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Origin": "https://cjcx.neea.edu.cn",
        "Referer": "https://cjcx.neea.edu.cn/"
    })

    print(f"\n[3/3] 开始批量查询...")
    results = []
    for i, student in enumerate(students):
        name = student["name"]
        id_num = student["id"]
        cls = student["class"]
        key = f"{cls}|{name}"
        progress = f"[{i+1}/{total}]"

        # 断点续跑:已查询过的直接复用
        if key in existing:
            results.append(existing[key])
            print(f"{progress} {cls} {name} 已查过,跳过")
            continue

        print(f"{progress} {cls} {name}", end=" ")

        entry = {
            "class": cls,
            "name": name,
            "id": mask_id(id_num),   # 脱敏后写入
            "id_short": len(id_num) < 15,
            "cet4": None,
            "cet6": None,
            "queried": False
        }

        # 身份证不完整,跳过查询
        if len(id_num) < 15:
            entry["status"] = "id_incomplete"
            entry["queried"] = True
            results.append(entry)
            print("跳过(证件号不完整)")
            continue

        # 查四级
        score4, cet4_confirmed = query_score(session, name, id_num, km=1)
        if score4:
            entry["cet4"] = score4
            print(f"四级:{score4['total']}", end=" ")
        time.sleep(REQUEST_INTERVAL)

        # 查六级
        score6, cet6_confirmed = query_score(session, name, id_num, km=2)
        if score6:
            entry["cet6"] = score6
            print(f"六级:{score6['total']}", end=" ")
        time.sleep(REQUEST_INTERVAL)

        entry["queried"] = cet4_confirmed and cet6_confirmed
        if not entry["queried"]:
            entry["status"] = "pending_retry"
            print("待下次重试")
        elif score4 or score6:
            entry["status"] = "found"
            print("OK")
        else:
            entry["status"] = "not_found"
            print("--")

        results.append(entry)

        # 每 50 人保存一次进度 + push 到仓库 + 冷却防限流
        if (i + 1) % 50 == 0:
            save_results(results)
            push_progress(i + 1, total)
            print(f"  --- 已保存进度 ({i+1}/{total})，冷却 10 秒 ---", flush=True)
            time.sleep(10)

    # 最终保存
    save_results(results)

    # 统计
    print("\n" + "=" * 50)
    print("  查询完成!")
    print("=" * 50)
    cet4_count = sum(1 for r in results if r["cet4"])
    cet6_count = sum(1 for r in results if r["cet6"])
    found_count = sum(1 for r in results if r["cet4"] or r["cet6"])
    print(f"  总人数:   {total}")
    print(f"  查到成绩: {found_count} 人(四级 {cet4_count} / 六级 {cet6_count})")
    print(f"  未查到:   {total - found_count} 人")
    print(f"\n  结果已保存到: {OUTPUT_FILE}")
    print(f"\n  将 scores.json 和 index.html 一起部署到 GitHub Pages 即可")


if __name__ == "__main__":
    main()
